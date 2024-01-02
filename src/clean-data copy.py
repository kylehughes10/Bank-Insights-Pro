import pandas as pd
import re
import os
import openpyxl  # Import openpyxl at the beginning of the script

# Define patterns for parsing
FDIC_CERT_PATTERN = r'FDIC Certificate # (\d+)'
DATE_PATTERN = r'\b(03/31|06/30|09/30|12/31)/(\d{4})\b'
PEER_GROUP_PATTERN = r'The current peer group for this bank is: (\d+)'
METRIC_PATTERN_REFINED = r'\n\s+(.*?)\s+(\-?\d+\.\d+)\s+(\-?\d+\.\d+)\s+(\d+)\s+'

# Global variable to store dates once extracted
global_dates = None

def read_ubpr_report(file_path):
    with open(file_path, 'r') as file:
        return file.read()

def extract_fdic_cert(report_content):
    match = re.search(FDIC_CERT_PATTERN, report_content)
    return match.group(1) if match else None

def extract_dates(report_content):
    global global_dates
    if global_dates is not None:
        print("Using cached dates")  # Debugging line
        return global_dates
    all_dates = re.findall(DATE_PATTERN, report_content)
    quarter_end_dates = [f'{month_year[0]}/{month_year[1]}' for month_year in all_dates if int(month_year[1]) > 2019]
    print(f"Found {len(quarter_end_dates)} quarter-end dates")  # Debugging line
    global_dates = quarter_end_dates
    return quarter_end_dates

def extract_peer_group(report_content):
    match = re.search(PEER_GROUP_PATTERN, report_content, re.MULTILINE)
    return match.group(1) if match else None

def extract_metrics_refined(report_content, dates):
    metrics_matches = re.findall(METRIC_PATTERN_REFINED, report_content)
    print(f"Found {len(metrics_matches)} metric matches")  # Debugging line
    structured_metrics = []
    for metric_match in metrics_matches:
        name = metric_match[0]
        values = re.findall(r'\d+\.\d+', metric_match[0] + metric_match[1])
        expected_values_count = len(dates) * 3
        if len(values) == expected_values_count:
            for i, date in enumerate(dates):
                bank_value = values[i*3]
                peer_group_value = values[i*3+1]
                percentile = values[i*3+2]
                structured_metrics.append({
                    'Date': date,
                    'Metric': name.strip(),
                    'Bank Value': bank_value,
                    'Peer Group Value': peer_group_value,
                    'Percentile': percentile
                })
        else:
            print(f"Warning: Metric '{name}' has {len(values)} values but expected {expected_values_count}. Skipping this metric.")
    return structured_metrics

def structure_data_refined(fdic_cert, dates, peer_group, metrics):
    data = []
    for metric in metrics:
        row = {
            'FDIC Certification #': fdic_cert,
            'Date': metric['Date'],
            'Peer Group': peer_group,
            f'{metric["Metric"]}': metric['Bank Value'],
            f'{metric["Metric"]} PG {peer_group}': metric['Peer Group Value'],
            f'{metric["Metric"]} PCT': metric['Percentile']
        }
        data.append(row)
    return pd.DataFrame(data)

def write_to_excel(data, excel_path):
    book = None
    if os.path.exists(excel_path):
        book = openpyxl.load_workbook(excel_path)  # This should now work with openpyxl imported

    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        writer.book = book
        if 'Structured UBPR Data' in writer.book.sheetnames:
            startrow = writer.book['Structured UBPR Data'].max_row
        else:
            startrow = None

        data.to_excel(writer, sheet_name='Structured UBPR Data', index=False, startrow=startrow)

if __name__ == "__main__":
    ubpr_report_path = "C:/Users/kyleh/OneDrive/Documents/YouTube/UBPR/ubpr-report-generator/public/UBPR Report.txt"
    excel_path = "C:/Users/kyleh/OneDrive/Documents/YouTube/UBPR/ubpr-report-generator/public/UBPR Example.xlsx"

    report_content = read_ubpr_report(ubpr_report_path)
    fdic_cert = extract_fdic_cert(report_content)
    dates = extract_dates(report_content)
    peer_group = extract_peer_group(report_content)
    metrics = extract_metrics_refined(report_content, dates)
    structured_data = structure_data_refined(fdic_cert, dates, peer_group, metrics)
    write_to_excel(structured_data, excel_path)
