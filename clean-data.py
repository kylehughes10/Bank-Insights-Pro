import pandas as pd
from openpyxl import load_workbook
import re

def parse_ubpr_report(report_path):
    with open(report_path, 'r', encoding='utf-8') as file:
        lines = file.readlines()

    data_collection = []
    current_metrics = {}
    dates = []

    # Define the pattern to match dates.
    date_pattern = re.compile(r"(\d{2}/\d{2}/\d{4})")

    # Adjust this index if the line number changes.
    date_line_index = 88 - 1  # Python lists are zero-indexed, so line 88 is at index 87.

    # Ensure the file has enough lines.
    if date_line_index < len(lines):
        line = lines[date_line_index]
        print(f"Checking for dates in line {date_line_index + 1}: {line.strip()}")  # Debugging

        # Find all dates in the line using the defined pattern.
        dates = date_pattern.findall(line)

        if dates:
            print(f"Found dates: {dates}")  # Debugging
        else:
            print("No dates found. Please check the date pattern and file content.")
            return pd.DataFrame()

    else:
        print(f"File doesn't have {date_line_index + 1} lines.")
        return pd.DataFrame()

    # Sort the dates.
    dates = sorted(set(dates), key=lambda x: pd.to_datetime(x, format='%m/%d/%Y'))

    for line in lines:
        data_match = data_pattern.search(line)
        if data_match:
            metric = data_match.group('metric').strip()
            values = data_match.group('values').split()
            
            expected_length = len(dates) * 3
            if len(values) != expected_length:
                print(f"Warning: The number of values for {metric} does not match the expected count. Found {len(values)}, expected {expected_length}.")
                continue

            for i, date in enumerate(dates):
                index = i * 3
                if date not in current_metrics:
                    current_metrics[date] = {'FDIC Certificate #': '4239', 'Date': date}
                try:
                    current_metrics[date][f"{metric} BANK"] = values[index]
                    current_metrics[date][f"{metric} PG2"] = values[index + 1]
                    current_metrics[date][f"{metric} PCT"] = values[index + 2]
                except IndexError:
                    print(f"Error: Index out of range for {metric} on date {date}.")
                    continue

    data_collection = [current_metrics[date] for date in dates]
    return pd.DataFrame(data_collection)

def add_sheet_to_excel(new_data_df, excel_file_path, sheet_name):
    try:
        book = load_workbook(excel_file_path)
    except Exception as e:
        print(f"Failed to load the existing Excel file: {e}")
        return

    if sheet_name in book.sheetnames:
        std = book[sheet_name]
        book.remove(std)

    try:
        book.save(excel_file_path)
    except Exception as e:
        print(f"Failed to save the Excel file: {e}")
        return

    try:
        with pd.ExcelWriter(excel_file_path, engine='openpyxl', mode='a') as writer:
            new_data_df.to_excel(writer, sheet_name=sheet_name, index=False)
    except Exception as e:
        print(f"Failed to write to the Excel file: {e}")

def main():
    # Replace with the actual path to your UBPR report file
    report_path = 'C:/Users/klhughes/AppData/Local/Programs/Python/Python312/Practice Python/Bank UBPR/UBPR Report.txt'
    # Replace with the actual path to your Excel file
    excel_file_path = 'C:/Users/klhughes/AppData/Local/Programs/Python/Python312/Practice Python/Bank UBPR/UBPR Example.xlsx'
    
    new_data_df = parse_ubpr_report(report_path)

    print("Data extracted to DataFrame:")
    print(new_data_df)
    
    if not new_data_df.empty:
        add_sheet_to_excel(new_data_df, excel_file_path, sheet_name='Parsed UBPR Data')
        print(f"Attempted to add new sheet to {excel_file_path}")
    else:
        print("No data to write to Excel.")

if __name__ == "__main__":
    main()